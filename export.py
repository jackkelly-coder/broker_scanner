import csv
import os
import sqlite3
from typing import Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import app_config
from database import DB_NAME

XLSX_PATH = os.path.join(app_config.EXPORT_DIR, "assignments.xlsx")
CSV_PATH = os.path.join(app_config.EXPORT_DIR, "assignments.csv")


def fetch_assignments() -> tuple[list[str], list[tuple]]:
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            title,
            company,
            location,
            location_bucket,
            published,
            url,
            scraped_at
        FROM assignments
        ORDER BY scraped_at DESC, company ASC, title ASC
        """
    )
    rows = cur.fetchall()
    conn.close()

    headers = [
        "title",
        "company",
        "location_raw",
        "location_bucket",
        "published",
        "url",
        "scraped_at",
    ]
    return headers, rows


def export_csv(headers: Sequence[str], rows: Sequence[Sequence]) -> None:
    os.makedirs(app_config.EXPORT_DIR, exist_ok=True)
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"Wrote CSV: {CSV_PATH}")


def export_xlsx(headers: Sequence[str], rows: Sequence[Sequence]) -> None:
    os.makedirs(app_config.EXPORT_DIR, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "assignments"

    worksheet.append(list(headers))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    url_column_index = headers.index("url") + 1

    for row_index, row in enumerate(rows, start=2):
        worksheet.append(list(row))
        url_value = row[url_column_index - 1]
        if url_value:
            cell = worksheet.cell(row=row_index, column=url_column_index)
            cell.hyperlink = url_value
            cell.value = url_value
            cell.style = "Hyperlink"

    for col_index in range(1, len(headers) + 1):
        max_len = 0
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=min(worksheet.max_row, 2000),
            min_col=col_index,
            max_col=col_index,
        ):
            value = row[0].value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_index)].width = min(max_len + 2, 60)

    workbook.save(XLSX_PATH)
    print(f"Wrote XLSX: {XLSX_PATH}")


def export_all() -> dict:
    headers, rows = fetch_assignments()
    export_csv(headers, rows)
    export_xlsx(headers, rows)
    print(f"Export complete. Rows: {len(rows)}")
    return {"rows": len(rows), "csv_path": CSV_PATH, "xlsx_path": XLSX_PATH}


if __name__ == "__main__":
    export_all()